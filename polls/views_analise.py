from django.shortcuts import render
from django.db.models import Avg, Count, F, Q, Max, Min
from django.utils import timezone
from datetime import timedelta
from .models import Crianca, Acao
import json
import numpy as np
from scipy.fft import fft, fftfreq
import math
import logging
from decimal import Decimal

# Novas importações para exportação
import pandas as pd
from io import BytesIO
from django.http import HttpResponse, JsonResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus import Image
import matplotlib.pyplot as plt
import io

logger = logging.getLogger(__name__)

# Função helper para converter Decimal para float
def decimal_to_float(obj):
    if isinstance(obj, Decimal):
        return float(obj)
    raise TypeError

# Custom JSON encoder para Decimal
class DecimalEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, Decimal):
            return float(obj)
        return super().default(obj)

def analise(request):
    criancas = Crianca.objects.all()
    crianca_id = request.GET.get("crianca")
    agrupar_por = request.GET.get("agrupar_por", "crianca")
    
    # Dados principais
    tempo_reacao_serie = []
    pontuacao_por_tipo = []
    distribuicao = []
    dados_espectrais = []
    estatisticas_gerais = {}
    crianca = None
    dados_jitter = []
    
    # Buscar todas as ações com tempo_reacao
    todas_acoes = Acao.objects.filter(tempo_reacao__isnull=False).select_related('crianca')
    
    if crianca_id:
        try:
            crianca = Crianca.objects.get(id=crianca_id)
            resultados = todas_acoes.filter(crianca_id=crianca_id).order_by("created_at")
            
            if resultados.exists():
                # 1. Gráfico de linha - Tempo de Reação por Tentativa
                tempo_reacao_serie = []
                for i, r in enumerate(resultados):
                    if r.tempo_reacao is not None:
                        tempo_reacao_serie.append({
                            "tentativa": i + 1,
                            "tempo_reacao": round(float(r.tempo_reacao), 3),
                            "marcador": r.marcador if hasattr(r, 'marcador') else None
                        })
                
                # 2. Gráfico de barras - Pontuação por Tipo de Ação
                # CORRIGIDO: usar 'tipo_acao' como alias, não 'tipo'
                pontuacao_por_tipo = list(
                    resultados.values(tipo_acao=F("tipo"))
                    .annotate(
                        pontuacao_media=Avg("pontuacao"),
                        quantidade=Count("id")
                    )
                    .order_by("-pontuacao_media")
                )
                
                tipo_map = {
                    'click': 'Clique',
                    'drag': 'Arrastar', 
                    'type': 'Digitar',
                    'select': 'Selecionar',
                    'submit': 'Enviar',
                    'next': 'Avançar',
                    'back': 'Voltar',
                    'hint': 'Pedir Dica',
                    'skip': 'Pular',
                }
                
                for item in pontuacao_por_tipo:
                    if item['tipo_acao'] in tipo_map:
                        item['tipo_acao'] = tipo_map[item['tipo_acao']]
                    if item['pontuacao_media'] is not None:
                        item['pontuacao_media'] = round(float(item['pontuacao_media']), 2)
                    else:
                        item['pontuacao_media'] = 0
                
                # 3. Distribuição por Faixa de Tempo
                total = resultados.count()
                distribuicao = [
                    {
                        "faixa": "0-2s",
                        "quantidade": resultados.filter(tempo_reacao__lt=2).count(),
                        "cor": "#22c55e",
                        "porcentagem": 0
                    },
                    {
                        "faixa": "2-4s", 
                        "quantidade": resultados.filter(tempo_reacao__gte=2, tempo_reacao__lt=4).count(),
                        "cor": "#3b82f6",
                        "porcentagem": 0
                    },
                    {
                        "faixa": "4-6s",
                        "quantidade": resultados.filter(tempo_reacao__gte=4, tempo_reacao__lt=6).count(),
                        "cor": "#f59e0b",
                        "porcentagem": 0
                    },
                    {
                        "faixa": "6+s",
                        "quantidade": resultados.filter(tempo_reacao__gte=6).count(),
                        "cor": "#ef4444",
                        "porcentagem": 0
                    },
                ]
                
                for item in distribuicao:
                    item["porcentagem"] = round((item["quantidade"] / total) * 100, 1) if total > 0 else 0
                
                # 4. Dados Espectrais
                tempos_reacao = list(resultados.filter(tempo_reacao__isnull=False).values_list('tempo_reacao', flat=True))
                tempos_reacao = [float(t) for t in tempos_reacao if t is not None]
                
                if len(tempos_reacao) > 3:
                    dados_espectrais = calcular_analise_espectral(tempos_reacao)
                
                # 5. Estatísticas Gerais
                tempo_medio = resultados.aggregate(Avg('tempo_reacao'))['tempo_reacao__avg']
                tempo_minimo = resultados.aggregate(Min('tempo_reacao'))['tempo_reacao__min']
                tempo_maximo = resultados.aggregate(Max('tempo_reacao'))['tempo_reacao__max']
                pontuacao_media = resultados.aggregate(Avg('pontuacao'))['pontuacao__avg']
                
                estatisticas_gerais = {
                    "total_acoes": total,
                    "tempo_medio": round(float(tempo_medio), 3) if tempo_medio is not None else 0,
                    "tempo_minimo": round(float(tempo_minimo), 3) if tempo_minimo is not None else 0,
                    "tempo_maximo": round(float(tempo_maximo), 3) if tempo_maximo is not None else 0,
                    "pontuacao_media": round(float(pontuacao_media), 2) if pontuacao_media is not None else 0,
                    "taxa_acertos": calcular_taxa_acertos(resultados),
                    "desvio_padrao": round(calcular_desvio_padrao(tempos_reacao), 3) if tempos_reacao else 0,
                    "data_inicio": resultados.first().created_at.strftime("%d/%m/%Y %H:%M") if resultados.first() else None,
                    "data_fim": resultados.last().created_at.strftime("%d/%m/%Y %H:%M") if resultados.last() else None,
                }
                
                # Dados para Jitter Plot
                dados_jitter = processar_dados_jitter(resultados, agrupar_por)
                
            else:
                estatisticas_gerais = {
                    "total_acoes": 0,
                    "tempo_medio": 0,
                    "tempo_minimo": 0,
                    "tempo_maximo": 0,
                    "pontuacao_media": 0,
                    "taxa_acertos": 0,
                    "desvio_padrao": 0,
                    "data_inicio": None,
                    "data_fim": None,
                }
                
        except Crianca.DoesNotExist:
            logger.warning("Crianca id=%s nao encontrada", crianca_id)
            crianca = None
            estatisticas_gerais = {}
        except Exception as e:
            logger.exception("Erro ao gerar analise para crianca_id=%s: %s", crianca_id, str(e))
            crianca = None
            estatisticas_gerais = {}
    else:
        # Todas as crianças - dados para Jitter Plot
        if todas_acoes.exists():
            dados_jitter = processar_dados_jitter(todas_acoes, agrupar_por)
            estatisticas_gerais = {
                "total_acoes": todas_acoes.count(),
                "tempo_medio": 0,
                "tempo_minimo": 0,
                "tempo_maximo": 0,
                "pontuacao_media": 0,
                "taxa_acertos": 0,
                "desvio_padrao": 0,
                "data_inicio": None,
                "data_fim": None,
            }
    
    contexto = {
        "criancas": criancas,
        "crianca_id_selecionada": crianca_id or "",
        "crianca_nome": crianca.nome if crianca else "",
        "tempo_reacao_serie_json": json.dumps(tempo_reacao_serie, cls=DecimalEncoder),
        "pontuacao_por_tipo_json": json.dumps(pontuacao_por_tipo, cls=DecimalEncoder),
        "distribuicao_json": json.dumps(distribuicao, cls=DecimalEncoder),
        "dados_espectrais_json": json.dumps(dados_espectrais, cls=DecimalEncoder),
        "estatisticas_gerais": estatisticas_gerais,
        "has_data": bool(tempo_reacao_serie),
        "dados_jitter_json": json.dumps(dados_jitter, cls=DecimalEncoder),
        "agrupar_por": agrupar_por,
    }
    
    return render(request, "polls/analise.html", contexto)

def processar_dados_jitter(resultados, agrupar_por="crianca"):
    """Processa dados para o gráfico Jitter Plot"""
    dados = []
    
    if agrupar_por == "crianca":
        for crianca in Crianca.objects.all():
            acoes = resultados.filter(crianca=crianca)
            for acao in acoes:
                dados.append({
                    'grupo': crianca.nome,
                    'valor': float(acao.tempo_reacao) if acao.tempo_reacao else 0,
                    'pontuacao': float(acao.pontuacao) if acao.pontuacao else 0,
                    'marcador': acao.marcador if hasattr(acao, 'marcador') else None,
                    'id': acao.id,
                })
    elif agrupar_por == "tipo":
        tipo_map = {
            'click': 'Clique', 'drag': 'Arrastar', 'type': 'Digitar',
            'select': 'Selecionar', 'submit': 'Enviar', 'next': 'Avançar',
            'back': 'Voltar', 'hint': 'Pedir Dica', 'skip': 'Pular',
        }
        for acao in resultados:
            grupo = tipo_map.get(acao.tipo, acao.tipo)
            dados.append({
                'grupo': grupo,
                'valor': float(acao.tempo_reacao) if acao.tempo_reacao else 0,
                'pontuacao': float(acao.pontuacao) if acao.pontuacao else 0,
                'marcador': acao.marcador if hasattr(acao, 'marcador') else None,
                'id': acao.id,
            })
    else:
        for acao in resultados:
            faixa = acao.crianca.faixa_etaria if hasattr(acao.crianca, 'faixa_etaria') else 'Não informada'
            dados.append({
                'grupo': faixa,
                'valor': float(acao.tempo_reacao) if acao.tempo_reacao else 0,
                'pontuacao': float(acao.pontuacao) if acao.pontuacao else 0,
                'marcador': acao.marcador if hasattr(acao, 'marcador') else None,
                'id': acao.id,
            })
    
    return dados

def calcular_analise_espectral(tempos_reacao):
    """Calcula a análise espectral usando FFT"""
    try:
        tempos_limpos = [float(t) for t in tempos_reacao if t is not None]
        n = len(tempos_limpos)
        
        if n < 10:
            return [
                {"tempo": i, "intensidade": min(100, t * 10), "frequencia": 0} 
                for i, t in enumerate(tempos_limpos[:30])
            ]
        
        yf = fft(tempos_limpos)
        xf = fftfreq(n, 1)[:n//2]
        magnitude = 2.0/n * np.abs(yf[:n//2])
        
        if len(magnitude) > 0 and np.max(magnitude) > 0:
            magnitude = (magnitude / np.max(magnitude)) * 100
        
        dados_espectrais = []
        for i in range(min(30, len(xf))):
            if i < len(magnitude):
                dados_espectrais.append({
                    "tempo": i,
                    "intensidade": round(float(magnitude[i]), 2),
                    "frequencia": round(float(xf[i]), 4)
                })
        
        return dados_espectrais
        
    except Exception as e:
        logger.exception("Erro na análise espectral: %s", str(e))
        tempos_limpos = [float(t) for t in tempos_reacao if t is not None]
        return [
            {"tempo": i, "intensidade": min(100, t * 10), "frequencia": 0} 
            for i, t in enumerate(tempos_limpos[:30])
        ]

def calcular_taxa_acertos(resultados):
    """Calcula a taxa de acertos baseada na pontuação"""
    try:
        acertos = resultados.filter(pontuacao__gt=0).count()
        total = resultados.count()
        if total > 0:
            return round((acertos / total) * 100, 1)
        return 0
    except:
        return 0

def calcular_desvio_padrao(dados):
    """Calcula o desvio padrão de uma lista de dados"""
    try:
        dados_limpos = [float(d) for d in dados if d is not None]
        if len(dados_limpos) < 2:
            return 0
        media = sum(dados_limpos) / len(dados_limpos)
        variancia = sum((x - media) ** 2 for x in dados_limpos) / len(dados_limpos)
        return math.sqrt(variancia)
    except:
        return 0

def api_dados_analise(request):
    """API endpoint para os componentes React obterem dados de análise"""
    crianca_id = request.GET.get("crianca_id")
    
    response_data = {
        "tempo_reacao_serie": [],
        "pontuacao_por_tipo": [],
        "distribuicao_tempo_resposta": [],
        "dados_espectrais": [],
        "total_acoes": 0
    }
    
    if crianca_id:
        try:
            resultados = Acao.objects.filter(
                crianca_id=crianca_id,
                tempo_reacao__isnull=False
            ).order_by("created_at")
            
            if resultados.exists():
                tempo_reacao_serie = []
                for i, r in enumerate(resultados):
                    if r.tempo_reacao is not None:
                        tempo_reacao_serie.append({
                            "tentativa": i + 1, 
                            "tempo_reacao": round(float(r.tempo_reacao), 3),
                            "marcador": r.marcador if hasattr(r, 'marcador') else None
                        })
                response_data["tempo_reacao_serie"] = tempo_reacao_serie
                
                tipo_map = {
                    'click': 'Clique', 
                    'drag': 'Arrastar', 
                    'type': 'Digitar',
                    'select': 'Selecionar',
                    'submit': 'Enviar',
                    'next': 'Avançar',
                    'back': 'Voltar',
                    'hint': 'Pedir Dica',
                    'skip': 'Pular',
                }
                
                # CORRIGIDO: usar 'tipo_acao' como alias
                pontuacao_por_tipo = list(
                    resultados.values(tipo_acao=F("tipo"))
                    .annotate(pontuacao_media=Avg("pontuacao"))
                )
                
                for item in pontuacao_por_tipo:
                    item["pontuacao_media"] = round(float(item["pontuacao_media"]), 2) if item["pontuacao_media"] is not None else 0
                    if item["tipo_acao"] in tipo_map:
                        item["tipo_acao"] = tipo_map[item["tipo_acao"]]
                
                response_data["pontuacao_por_tipo"] = pontuacao_por_tipo
                
                total = resultados.count()
                response_data["distribuicao_tempo_resposta"] = [
                    {"faixa": "0-2s", "quantidade": resultados.filter(tempo_reacao__lt=2).count()},
                    {"faixa": "2-4s", "quantidade": resultados.filter(tempo_reacao__gte=2, tempo_reacao__lt=4).count()},
                    {"faixa": "4-6s", "quantidade": resultados.filter(tempo_reacao__gte=4, tempo_reacao__lt=6).count()},
                    {"faixa": "6+s", "quantidade": resultados.filter(tempo_reacao__gte=6).count()},
                ]
                
                tempos = list(resultados.filter(tempo_reacao__isnull=False).values_list('tempo_reacao', flat=True))
                tempos = [float(t) for t in tempos if t is not None]
                if tempos:
                    response_data["dados_espectrais"] = calcular_analise_espectral(tempos)
                
                response_data["total_acoes"] = total
                
        except Exception as e:
            logger.exception("Erro na API de análise: %s", str(e))
            response_data["erro"] = str(e)
    
    return JsonResponse(response_data)


# ============================================================
# EXPORTAÇÃO EXCEL
# ============================================================

def exportar_excel(request):
    """Exporta dados de análise para Excel (.xlsx) com formatação profissional"""
    
    crianca_id = request.GET.get("crianca_id")
    crianca_nome = "Todos"
    
    queryset = Acao.objects.filter(tempo_reacao__isnull=False).select_related('crianca')
    
    if crianca_id:
        queryset = queryset.filter(crianca_id=crianca_id)
        try:
            crianca = Crianca.objects.get(id=crianca_id)
            crianca_nome = crianca.nome
        except Crianca.DoesNotExist:
            pass
    
    queryset = queryset.order_by("created_at")
    
    if not queryset.exists():
        return HttpResponse("Nenhum dado disponível para exportar", status=404)
    
    wb = Workbook()
    
    # Folha 1: Dados Brutos
    ws1 = wb.active
    ws1.title = "Dados Brutos"
    
    headers = ["ID", "Criança", "Tipo", "Pontuação", "Tempo de Reação (s)", 
               "Marcador", "Timestamp", "Data", "Hora"]
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    for row_idx, acao in enumerate(queryset, 2):
        ws1.cell(row=row_idx, column=1, value=acao.id)
        ws1.cell(row=row_idx, column=2, value=acao.crianca.nome if acao.crianca else "N/A")
        ws1.cell(row=row_idx, column=3, value=acao.tipo)
        ws1.cell(row=row_idx, column=4, value=float(acao.pontuacao) if acao.pontuacao else 0)
        ws1.cell(row=row_idx, column=5, value=float(acao.tempo_reacao) if acao.tempo_reacao else None)
        ws1.cell(row=row_idx, column=6, value=acao.marcador if hasattr(acao, 'marcador') and acao.marcador else "")
        ws1.cell(row=row_idx, column=7, value=acao.created_at.strftime("%Y-%m-%d %H:%M:%S"))
        ws1.cell(row=row_idx, column=8, value=acao.created_at.strftime("%Y-%m-%d"))
        ws1.cell(row=row_idx, column=9, value=acao.created_at.strftime("%H:%M:%S"))
        
        for col in range(1, 10):
            ws1.cell(row=row_idx, column=col).border = border
    
    for col in range(1, 10):
        ws1.column_dimensions[chr(64 + col)].width = 18
    
    # Folha 2: Estatísticas
    ws2 = wb.create_sheet("Estatísticas")
    
    tempos = [float(a.tempo_reacao) for a in queryset if a.tempo_reacao is not None]
    pontuacoes = [float(a.pontuacao) for a in queryset if a.pontuacao is not None]
    
    stats_data = [
        ["Métrica", "Valor"],
        ["Total de Ações", queryset.count()],
        ["Tempo Médio (s)", f"{np.mean(tempos):.3f}" if tempos else "N/A"],
        ["Tempo Mínimo (s)", f"{np.min(tempos):.3f}" if tempos else "N/A"],
        ["Tempo Máximo (s)", f"{np.max(tempos):.3f}" if tempos else "N/A"],
        ["Desvio Padrão (s)", f"{np.std(tempos):.3f}" if tempos else "N/A"],
        ["Pontuação Média", f"{np.mean(pontuacoes):.2f}" if pontuacoes else "N/A"],
        ["Taxa de Acertos", f"{(sum(1 for p in pontuacoes if p > 0) / len(pontuacoes) * 100):.1f}%" if pontuacoes else "N/A"],
    ]
    
    for row_idx, row_data in enumerate(stats_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            cell.border = border
    
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 20
    
    # Folha 3: Distribuição por Tempo
    ws3 = wb.create_sheet("Distribuição por Tempo")
    
    faixas = [
        ("0-2s", queryset.filter(tempo_reacao__lt=2).count()),
        ("2-4s", queryset.filter(tempo_reacao__gte=2, tempo_reacao__lt=4).count()),
        ("4-6s", queryset.filter(tempo_reacao__gte=4, tempo_reacao__lt=6).count()),
        ("6+s", queryset.filter(tempo_reacao__gte=6).count()),
    ]
    total = queryset.count()
    
    distrib_data = [["Faixa", "Quantidade", "Porcentagem"]]
    for faixa, qtd in faixas:
        distrib_data.append([faixa, qtd, f"{(qtd / total * 100):.1f}%" if total > 0 else "0%"])
    
    for row_idx, row_data in enumerate(distrib_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            cell.border = border
    
    for col in range(1, 4):
        ws3.column_dimensions[chr(64 + col)].width = 18
    
    # Folha 4: Pontuação por Tipo
    ws4 = wb.create_sheet("Pontuação por Tipo")
    
    tipo_map = {
        'click': 'Clique', 'drag': 'Arrastar', 'type': 'Digitar',
        'select': 'Selecionar', 'submit': 'Enviar', 'next': 'Avançar',
        'back': 'Voltar', 'hint': 'Pedir Dica', 'skip': 'Pular',
    }
    
    # CORRIGIDO: usar 'tipo_acao' como alias
    tipos_data = list(
        queryset.values(tipo_acao=F("tipo"))
        .annotate(media=Avg("pontuacao"), total=Count("id"))
        .order_by("-media")
    )
    
    tipo_stats = [["Tipo", "Pontuação Média", "Quantidade"]]
    for item in tipos_data:
        nome = tipo_map.get(item['tipo_acao'], item['tipo_acao'])
        tipo_stats.append([
            nome,
            f"{float(item['media']):.2f}" if item['media'] else "0",
            item['total']
        ])
    
    for row_idx, row_data in enumerate(tipo_stats, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws4.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            cell.border = border
    
    for col in range(1, 4):
        ws4.column_dimensions[chr(64 + col)].width = 20
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    nome_arquivo = f"analise_{crianca_nome.replace(' ', '_')}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'
    
    wb.save(response)
    return response


# ============================================================
# EXPORTAÇÃO PDF
# ============================================================

def exportar_pdf(request):
    """Exporta dados de análise para PDF com todos os gráficos (Jitter, Espectral, Pontuação)"""
    
    crianca_id = request.GET.get("crianca_id")
    agrupar_por = request.GET.get("agrupar_por", "crianca")
    crianca_nome = "Todos"
    
    queryset = Acao.objects.filter(tempo_reacao__isnull=False).select_related('crianca')
    
    if crianca_id:
        queryset = queryset.filter(crianca_id=crianca_id)
        try:
            crianca = Crianca.objects.get(id=crianca_id)
            crianca_nome = crianca.nome
        except Crianca.DoesNotExist:
            pass
    
    queryset = queryset.order_by("created_at")
    
    if not queryset.exists():
        return HttpResponse("Nenhum dado disponível para exportar", status=404)
    
    response = HttpResponse(content_type='application/pdf')
    nome_arquivo = f"analise_{crianca_nome.replace(' ', '_')}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'
    
    # Configurar documento com margens menores para caber mais conteúdo
    doc = SimpleDocTemplate(response, pagesize=A4, 
                           rightMargin=1.5*cm, leftMargin=1.5*cm,
                           topMargin=1.5*cm, bottomMargin=1.5*cm)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        alignment=TA_CENTER,
        spaceAfter=16
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=10,
        spaceBefore=10
    )
    
    # Coletar dados
    tempos = [float(a.tempo_reacao) for a in queryset if a.tempo_reacao is not None]
    pontuacoes = [float(a.pontuacao) for a in queryset if a.pontuacao is not None]
    
    # Dados para Jitter Plot
    dados_jitter = processar_dados_jitter(queryset, agrupar_por)
    
    # Dados para pontuação por tipo
    tipo_map = {
        'click': 'Clique', 'drag': 'Arrastar', 'type': 'Digitar',
        'select': 'Selecionar', 'submit': 'Enviar', 'next': 'Avançar',
        'back': 'Voltar', 'hint': 'Pedir Dica', 'skip': 'Pular',
    }
    tipos_data = list(
        queryset.values(tipo_acao=F("tipo"))
        .annotate(media=Avg("pontuacao"), total=Count("id"))
        .order_by("-media")
    )
    
    # Dados espectrais
    dados_espectrais = calcular_analise_espectral(tempos) if len(tempos) > 3 else []
    
    story = []
    
    # ============================================================
    # TÍTULO E ESTATÍSTICAS
    # ============================================================
    story.append(Paragraph(f"Relatório de Análise de Desempenho", title_style))
    story.append(Paragraph(f"Criança: {crianca_nome}", styles['Normal']))
    story.append(Paragraph(f"Data: {timezone.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    story.append(Spacer(1, 15))
    
    stats_data = [
        ['Métrica', 'Valor'],
        ['Total de Ações', str(queryset.count())],
        ['Tempo Médio (s)', f"{np.mean(tempos):.3f}" if tempos else "N/A"],
        ['Tempo Mínimo (s)', f"{np.min(tempos):.3f}" if tempos else "N/A"],
        ['Tempo Máximo (s)', f"{np.max(tempos):.3f}" if tempos else "N/A"],
        ['Desvio Padrão (s)', f"{np.std(tempos):.3f}" if tempos else "N/A"],
        ['Pontuação Média', f"{np.mean(pontuacoes):.2f}" if pontuacoes else "N/A"],
    ]
    
    table = Table(stats_data, colWidths=[6*cm, 6*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(table)
    story.append(Spacer(1, 20))
    
    # ============================================================
    # GRÁFICO 1: JITTER PLOT
    # ============================================================
    if dados_jitter and len(dados_jitter) > 0:
        story.append(Paragraph("1. Análise em Grupo - Jitter Plot", heading_style))
        
        try:
            fig1, ax1 = plt.subplots(figsize=(10, 6))
            
            # Agrupar dados
            grupos = {}
            for item in dados_jitter:
                chave = item['grupo'] or 'Desconhecido'
                if chave not in grupos:
                    grupos[chave] = []
                grupos[chave].append(item['valor'])
            
            cores = ['#8b5cf6', '#3b82f6', '#22c55e', '#f59e0b', '#ef4444', '#ec4899', '#06b6d4', '#84cc16']
            lista_grupos = list(grupos.keys())
            
            # Criar Jitter Plot
            import random
            random.seed(42)  # Para reprodutibilidade
            
            for idx, (nome_grupo, valores) in enumerate(grupos.items()):
                x_pos = idx + 1
                # Adicionar jitter
                x_jitter = [x_pos + (random.random() - 0.5) * 0.4 for _ in valores]
                cor = cores[idx % len(cores)]
                
                ax1.scatter(x_jitter, valores, alpha=0.6, s=40, 
                           color=cor, label=f'{nome_grupo} (n={len(valores)})')
                
                # Linha da média
                media = np.mean(valores)
                ax1.axhline(y=media, xmin=(x_pos-0.4)/len(lista_grupos), 
                           xmax=(x_pos+0.4)/len(lista_grupos), 
                           color=cor, linestyle='--', linewidth=2)
                ax1.text(x_pos + 0.3, media, f'μ={media:.2f}s', fontsize=8, color=cor)
            
            ax1.set_xlabel('Grupos')
            ax1.set_ylabel('Tempo de Reação (s)')
            ax1.set_title('Distribuição dos Tempos de Reação por Grupo')
            ax1.set_xticks(range(1, len(lista_grupos) + 1))
            ax1.set_xticklabels(lista_grupos, rotation=15, ha='right', fontsize=8)
            ax1.legend(loc='upper right', fontsize=8)
            ax1.grid(axis='y', alpha=0.3)
            ax1.set_ylim(0, max([max(v) for v in grupos.values()]) * 1.1 if grupos else 10)
            
            plt.tight_layout()
            
            img_buffer1 = io.BytesIO()
            plt.savefig(img_buffer1, format='png', dpi=150, bbox_inches='tight')
            plt.close()
            img_buffer1.seek(0)
            
            img1 = Image(img_buffer1, width=16*cm, height=8*cm)
            story.append(img1)
            story.append(Spacer(1, 10))
            
        except Exception as e:
            logger.exception("Erro ao gerar Jitter Plot para PDF: %s", str(e))
            story.append(Paragraph(f"Erro ao gerar gráfico Jitter: {str(e)}", styles['Normal']))
            story.append(Spacer(1, 10))
    
    # ============================================================
    # GRÁFICO 2: ANÁLISE ESPECTRAL
    # ============================================================
    if dados_espectrais and len(dados_espectrais) > 0:
        story.append(Paragraph("2. Análise Espectral (Fourier)", heading_style))
        
        try:
            fig2, ax2 = plt.subplots(figsize=(10, 4))
            
            intensidades = [item['intensidade'] for item in dados_espectrais]
            frequencias = [item['frequencia'] for item in dados_espectrais]
            
            ax2.plot(range(len(intensidades)), intensidades, 
                    color='#06b6d4', linewidth=2)
            ax2.fill_between(range(len(intensidades)), intensidades, 
                           alpha=0.3, color='#06b6d4')
            ax2.set_xlabel('Componente Espectral')
            ax2.set_ylabel('Intensidade (%)')
            ax2.set_title('Análise de Frequência dos Padrões de Resposta')
            ax2.grid(alpha=0.3)
            ax2.set_ylim(0, 105)
            
            plt.tight_layout()
            
            img_buffer2 = io.BytesIO()
            plt.savefig(img_buffer2, format='png', dpi=150, bbox_inches='tight')
            plt.close()
            img_buffer2.seek(0)
            
            img2 = Image(img_buffer2, width=16*cm, height=6*cm)
            story.append(img2)
            story.append(Spacer(1, 10))
            
        except Exception as e:
            logger.exception("Erro ao gerar gráfico espectral para PDF: %s", str(e))
            story.append(Paragraph(f"Erro ao gerar gráfico espectral: {str(e)}", styles['Normal']))
            story.append(Spacer(1, 10))
    
    # ============================================================
    # GRÁFICO 3: PONTUAÇÃO POR TIPO
    # ============================================================
    if tipos_data:
        story.append(Paragraph("3. Pontuação Média por Tipo de Ação", heading_style))
        
        try:
            fig3, ax3 = plt.subplots(figsize=(10, 4))
            
            nomes = []
            medias = []
            quantidades = []
            cores_barras = ['#22c55e', '#3b82f6', '#f59e0b', '#ef4444', '#8b5cf6', '#ec4899']
            
            for i, item in enumerate(tipos_data):
                nome = tipo_map.get(item['tipo_acao'], item['tipo_acao'])
                nomes.append(nome)
                medias.append(float(item['media']) if item['media'] else 0)
                quantidades.append(item['total'])
            
            bars = ax3.bar(nomes, medias, color=cores_barras[:len(nomes)], 
                          edgecolor='white', linewidth=1)
            ax3.set_ylabel('Pontuação Média')
            ax3.set_xlabel('Tipo de Ação')
            ax3.set_title('Pontuação Média por Tipo de Ação')
            ax3.grid(axis='y', alpha=0.3)
            ax3.set_ylim(0, max(medias) * 1.2 if medias else 10)
            
            # Adicionar valores e quantidades nas barras
            for bar, media, qtd in zip(bars, medias, quantidades):
                height = bar.get_height()
                ax3.text(bar.get_x() + bar.get_width()/2., height + 0.5,
                        f'{media:.1f} pts\n(n={qtd})',
                        ha='center', va='bottom', fontsize=8)
            
            plt.tight_layout()
            
            img_buffer3 = io.BytesIO()
            plt.savefig(img_buffer3, format='png', dpi=150, bbox_inches='tight')
            plt.close()
            img_buffer3.seek(0)
            
            img3 = Image(img_buffer3, width=16*cm, height=6*cm)
            story.append(img3)
            story.append(Spacer(1, 10))
            
        except Exception as e:
            logger.exception("Erro ao gerar gráfico de pontuação para PDF: %s", str(e))
            story.append(Paragraph(f"Erro ao gerar gráfico de pontuação: {str(e)}", styles['Normal']))
            story.append(Spacer(1, 10))
    
    # ============================================================
    # GRÁFICO 4: DISTRIBUIÇÃO POR TEMPO
    # ============================================================
    story.append(Paragraph("4. Distribuição por Tempo de Resposta", heading_style))
    
    try:
        fig4, ax4 = plt.subplots(figsize=(10, 4))
        faixas = ['0-2s', '2-4s', '4-6s', '6+s']
        valores = [
            queryset.filter(tempo_reacao__lt=2).count(),
            queryset.filter(tempo_reacao__gte=2, tempo_reacao__lt=4).count(),
            queryset.filter(tempo_reacao__gte=4, tempo_reacao__lt=6).count(),
            queryset.filter(tempo_reacao__gte=6).count(),
        ]
        cores4 = ['#22c55e', '#3b82f6', '#f59e0b', '#ef4444']
        
        bars4 = ax4.bar(faixas, valores, color=cores4, edgecolor='white', linewidth=2)
        ax4.set_ylabel('Quantidade')
        ax4.set_xlabel('Faixa de Tempo')
        ax4.set_title('Distribuição por Tempo de Resposta')
        ax4.grid(axis='y', alpha=0.3)
        
        total = queryset.count()
        for bar, valor in zip(bars4, valores):
            height = bar.get_height()
            porcentagem = (valor / total * 100) if total > 0 else 0
            ax4.text(bar.get_x() + bar.get_width()/2., height + 0.5,
                    f'{int(valor)}\n({porcentagem:.1f}%)',
                    ha='center', va='bottom', fontsize=9)
        
        plt.tight_layout()
        
        img_buffer4 = io.BytesIO()
        plt.savefig(img_buffer4, format='png', dpi=150, bbox_inches='tight')
        plt.close()
        img_buffer4.seek(0)
        
        img4 = Image(img_buffer4, width=16*cm, height=6*cm)
        story.append(img4)
        story.append(Spacer(1, 10))
        
    except Exception as e:
        logger.exception("Erro ao gerar gráfico de distribuição para PDF: %s", str(e))
        story.append(Paragraph(f"Erro ao gerar gráfico de distribuição: {str(e)}", styles['Normal']))
        story.append(Spacer(1, 10))
    
    # ============================================================
    # GRÁFICO 5: EVOLUÇÃO DO TEMPO DE REAÇÃO
    # ============================================================
    if len(tempos) > 1:
        story.append(Paragraph("5. Evolução do Tempo de Reação", heading_style))
        
        try:
            fig5, ax5 = plt.subplots(figsize=(10, 4))
            ax5.plot(range(1, len(tempos)+1), tempos, 
                    color='#8b5cf6', linewidth=2, marker='o', markersize=3)
            ax5.axhline(y=np.mean(tempos), color='red', linestyle='--', 
                       label=f'Média: {np.mean(tempos):.2f}s')
            ax5.set_xlabel('Tentativa')
            ax5.set_ylabel('Tempo de Reação (s)')
            ax5.set_title('Evolução do Tempo de Reação')
            ax5.legend(fontsize=8)
            ax5.grid(alpha=0.3)
            
            plt.tight_layout()
            
            img_buffer5 = io.BytesIO()
            plt.savefig(img_buffer5, format='png', dpi=150, bbox_inches='tight')
            plt.close()
            img_buffer5.seek(0)
            
            img5 = Image(img_buffer5, width=16*cm, height=6*cm)
            story.append(img5)
            
        except Exception as e:
            logger.exception("Erro ao gerar gráfico de evolução para PDF: %s", str(e))
            story.append(Paragraph(f"Erro ao gerar gráfico de evolução: {str(e)}", styles['Normal']))
    
    # ============================================================
    # RODAPÉ
    # ============================================================
    story.append(Spacer(1, 15))
    story.append(Paragraph(
        f"Relatório gerado em {timezone.now().strftime('%d/%m/%Y às %H:%M')} | Página 1 de 1",
        styles['Normal']
    ))
    
    # Construir PDF
    doc.build(story)
    return response


# ============================================================
# API DADOS GRUPO (JITTER PLOT)
# ============================================================

def api_dados_grupo(request):
    """API para dados de análise em grupo com Jitter Plot"""
    
    agrupar_por = request.GET.get("agrupar_por", "crianca")
    queryset = Acao.objects.filter(tempo_reacao__isnull=False).select_related('crianca')
    
    dados = []
    
    if agrupar_por == "crianca":
        for crianca in Crianca.objects.all():
            acoes = queryset.filter(crianca=crianca)
            if acoes.exists():
                for acao in acoes:
                    dados.append({
                        'id': acao.id,
                        'crianca_id': crianca.id,
                        'crianca_nome': crianca.nome,
                        'crianca': crianca.nome,
                        'tipo': acao.tipo,
                        'tempo_reacao': float(acao.tempo_reacao) if acao.tempo_reacao else 0,
                        'pontuacao': float(acao.pontuacao) if acao.pontuacao else 0,
                        'marcador': acao.marcador if hasattr(acao, 'marcador') else None,
                        'faixa_etaria': crianca.faixa_etaria if hasattr(crianca, 'faixa_etaria') else None,
                    })
    elif agrupar_por == "tipo":
        for acao in queryset:
            dados.append({
                'id': acao.id,
                'crianca_id': acao.crianca_id,
                'crianca_nome': acao.crianca.nome if acao.crianca else None,
                'crianca': acao.crianca.nome if acao.crianca else None,
                'tipo': acao.tipo,
                'tempo_reacao': float(acao.tempo_reacao) if acao.tempo_reacao else 0,
                'pontuacao': float(acao.pontuacao) if acao.pontuacao else 0,
                'marcador': acao.marcador if hasattr(acao, 'marcador') else None,
            })
    else:
        for crianca in Crianca.objects.all():
            faixa = crianca.faixa_etaria if hasattr(crianca, 'faixa_etaria') else 'Não informada'
            acoes = queryset.filter(crianca=crianca)
            for acao in acoes:
                dados.append({
                    'id': acao.id,
                    'crianca_id': crianca.id,
                    'crianca_nome': crianca.nome,
                    'crianca': crianca.nome,
                    'tipo': acao.tipo,
                    'tempo_reacao': float(acao.tempo_reacao) if acao.tempo_reacao else 0,
                    'pontuacao': float(acao.pontuacao) if acao.pontuacao else 0,
                    'marcador': acao.marcador if hasattr(acao, 'marcador') else None,
                    'faixa_etaria': faixa,
                })
    
    return JsonResponse({
        'dados': dados,
        'total': len(dados),
        'agrupado_por': agrupar_por
    })